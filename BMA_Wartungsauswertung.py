
#!/usr/bin/env python3
"""
BMA-Wartungsauswertung für ODS-, XLSX- und XLS-Prüflisten

Funktionen:
- Liest Hekatron-Ereignisspeicher als Excel-Export (.xlsx)
- Liest Esser-Ereignisspeicher als PDF
- Hekatron: „Revisionsalarm / Rauchalarm“ anhand von Nr und SubNr
- Esser: Ereignis „FEUER“; optional zusätzlich „VORALARM“ und „VORALARM-ENDE"
- Sucht die Meldegruppe in Spalte A der ODS- oder Excel-Prüfliste
- Spalte F entspricht Melder 1, G entspricht Melder 2 usw.
- Trägt die ausgewählte Quartalsnummer 1–4 ein
- Überschreibt keine bereits belegten Zellen
- Verwendet openpyxl für eine zuverlässige Windows-/Excel-Ausgabe
- Erstellt optional einen TXT-Prüfbericht mit nicht geprüften Meldern und unvollständigen Gruppen
- Kompakte Statistikzeile mit Quartalsstatus und Ampelanzeige

Hinweis:
Das Programm erstellt immer eine neue Prüfliste im Format der Vorlage. Die Originalvorlage bleibt unverändert.
"""

from __future__ import annotations

import copy
import json
import os
import re
import tkinter as tk
import xml.etree.ElementTree as ET
import zipfile
import traceback
import threading
import urllib.request
import urllib.error
import ssl
import webbrowser
import hashlib
import sys
import tempfile
from pathlib import Path
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any, Dict, List, Optional, Set, Tuple


try:
    import bcrypt
except ImportError:
    bcrypt = None

try:
    import certifi
except ImportError:
    certifi = None

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    load_workbook = None

    class InvalidFileException(Exception):
        pass

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

try:
    import xlrd
    from xlutils.copy import copy as copy_xls_workbook
except ImportError:
    xlrd = None
    copy_xls_workbook = None


# ----------------------------- Namensräume -----------------------------

NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    "style": "urn:oasis:names:tc:opendocument:xmlns:style:1.0",
    "fo": "urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0",
    "calcext": "urn:org:documentfoundation:names:experimental:calc:xmlns:calcext:1.0",
}

for prefix, uri in NS.items():
    ET.register_namespace(prefix, uri)

TABLE_ROW = f"{{{NS['table']}}}table-row"
TABLE_CELL = f"{{{NS['table']}}}table-cell"
COVERED_CELL = f"{{{NS['table']}}}covered-table-cell"
TEXT_P = f"{{{NS['text']}}}p"

ATTR_COL_REPEAT = f"{{{NS['table']}}}number-columns-repeated"
ATTR_ROW_REPEAT = f"{{{NS['table']}}}number-rows-repeated"
ATTR_STYLE_NAME = f"{{{NS['table']}}}style-name"
ATTR_FORMULA = f"{{{NS['table']}}}formula"
ATTR_VALUE_TYPE = f"{{{NS['office']}}}value-type"
ATTR_VALUE = f"{{{NS['office']}}}value"
ATTR_STRING_VALUE = f"{{{NS['office']}}}string-value"

# Vorlage: A = Meldegruppe, D = Anzahl Melder, F = Melder 1
MG_COLUMN_INDEX = 1
COUNT_COLUMN_INDEX = 4
FIRST_DETECTOR_COLUMN_INDEX = 6

VERSION = "1.0.1"
USERS_FILENAME = "users.json"
VERSION_URL = "https://raw.githubusercontent.com/tmk68htskp-bot/BMA-Wartungsauswertung/main/version.json"

MAX_LOGO_WIDTH = 180
MAX_LOGO_HEIGHT = 80


class Auswertungsfehler(Exception):
    """Fehler, der verständlich in der Bedienoberfläche angezeigt wird."""


def users_file_path() -> Path:
    """Liefert den Pfad zur lokalen, offline nutzbaren Benutzerdatei."""
    return Path(__file__).resolve().parent / USERS_FILENAME


def load_users() -> Dict[str, Dict[str, Any]]:
    """Lädt die lokale Benutzerdatei; Internet ist nicht erforderlich."""
    path = users_file_path()
    if not path.exists():
        raise Auswertungsfehler(
            f"Die Benutzerdatei '{USERS_FILENAME}' fehlt. "
            "Bitte lege sie in denselben Ordner wie das Programm."
        )

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise Auswertungsfehler(
            f"Die Benutzerdatei '{USERS_FILENAME}' konnte nicht gelesen werden: {exc}"
        ) from exc

    if not isinstance(data, dict):
        raise Auswertungsfehler(
            f"Die Benutzerdatei '{USERS_FILENAME}' hat ein ungültiges Format."
        )
    return data


def verify_user(username: str, password: str) -> Optional[str]:
    """Prüft ein Klartextpasswort gegen den gespeicherten bcrypt-Hash."""
    if bcrypt is None:
        raise Auswertungsfehler(
            "Die Bibliothek 'bcrypt' fehlt. Bitte ausführen: python -m pip install bcrypt"
        )

    users = load_users()
    record = users.get(username)
    if not isinstance(record, dict) or not record.get("aktiv", True):
        return None

    password_hash = record.get("password_hash") or record.get("password")
    if not isinstance(password_hash, str) or not password_hash.startswith("$2"):
        return None

    try:
        valid = bcrypt.checkpw(
            password.encode("utf-8"),
            password_hash.encode("utf-8"),
        )
    except (ValueError, TypeError):
        return None

    if valid:
        return str(record.get("rolle", "Benutzer"))
    return None


def parse_version(version: str) -> Tuple[int, ...]:
    """Wandelt z. B. 0.1.10 in (0, 1, 10) um."""
    cleaned = str(version).strip().lower().lstrip("v")
    parts = cleaned.split(".")
    if not parts or any(not part.isdigit() for part in parts):
        raise ValueError(f"Ungültige Versionsnummer: {version}")
    return tuple(int(part) for part in parts)


def create_ssl_context() -> ssl.SSLContext:
    """Erstellt einen geprüften HTTPS-Kontext, bevorzugt mit certifi."""
    if certifi is not None:
        return ssl.create_default_context(cafile=certifi.where())
    return ssl.create_default_context()


def load_online_version() -> Dict[str, Any]:
    """Lädt die öffentliche version.json sicher von GitHub."""
    request = urllib.request.Request(
        VERSION_URL,
        headers={
            "User-Agent": "BMA-Wartungsauswertung",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        },
    )
    try:
        with urllib.request.urlopen(
            request,
            timeout=8,
            context=create_ssl_context(),
        ) as response:
            raw = response.read().decode("utf-8-sig")
    except urllib.error.URLError as exc:
        reason = getattr(exc, "reason", exc)
        if isinstance(reason, ssl.SSLCertVerificationError):
            raise Auswertungsfehler(
                "Die HTTPS-Zertifikate konnten nicht geprüft werden. "
                "Bitte installiere 'certifi' mit: python3 -m pip install certifi"
            ) from exc
        raise

    data = json.loads(raw)
    if not isinstance(data, dict) or not data.get("version"):
        raise ValueError("Die Online-Versionsdatei ist ungültig.")
    return data


# ----------------------------- Hilfsfunktionen -----------------------------

def as_int(value: object) -> Optional[int]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    match = re.search(r"-?\d+(?:[.,]\d+)?", text)
    if not match:
        return None

    try:
        return int(float(match.group(0).replace(",", ".")))
    except ValueError:
        return None


def cell_text(cell: ET.Element) -> str:
    """Liest sichtbaren Text oder numerischen Zellwert aus einer ODS-Zelle."""
    paragraphs = cell.findall(f".//{TEXT_P}")
    if paragraphs:
        parts: List[str] = []
        for paragraph in paragraphs:
            parts.append("".join(paragraph.itertext()))
        text = "\n".join(parts).strip()
        if text:
            return text

    value = cell.get(ATTR_VALUE)
    if value is not None:
        return value

    string_value = cell.get(ATTR_STRING_VALUE)
    if string_value is not None:
        return string_value

    return ""


def repeated_count(element: ET.Element, attribute: str) -> int:
    try:
        return max(1, int(element.get(attribute, "1")))
    except ValueError:
        return 1


def clone_without_repeat(element: ET.Element) -> ET.Element:
    cloned = copy.deepcopy(element)
    cloned.attrib.pop(ATTR_COL_REPEAT, None)
    cloned.attrib.pop(ATTR_ROW_REPEAT, None)
    return cloned


def set_numeric_cell_value(cell: ET.Element, number: int) -> None:
    """Setzt einen ODS-Zellwert als Zahl und erhält den vorhandenen Zellstil."""
    style_name = cell.get(ATTR_STYLE_NAME)

    # Zellinhalt und eventuelle Formel entfernen.
    cell.clear()

    if style_name:
        cell.set(ATTR_STYLE_NAME, style_name)

    cell.set(ATTR_VALUE_TYPE, "float")
    cell.set(ATTR_VALUE, str(number))

    paragraph = ET.SubElement(cell, TEXT_P)
    paragraph.text = str(number)


def is_cell_element(element: ET.Element) -> bool:
    return element.tag in (TABLE_CELL, COVERED_CELL)


def get_cell_at(row: ET.Element, logical_column: int) -> Optional[ET.Element]:
    """
    Liefert die Zelle an einer logischen Spaltenposition.
    Berücksichtigt table:number-columns-repeated.
    """
    column = 1

    for child in list(row):
        if not is_cell_element(child):
            continue

        repeat = repeated_count(child, ATTR_COL_REPEAT)
        if column <= logical_column < column + repeat:
            return child
        column += repeat

    return None


def ensure_cell_at(row: ET.Element, logical_column: int) -> ET.Element:
    """
    Liefert eine einzelne, bearbeitbare Zelle an der gewünschten Spalte.
    Falls die Position in einer wiederholten Zelle liegt, wird diese sauber aufgeteilt.
    """
    children = list(row)
    logical_pos = 1

    for xml_index, child in enumerate(children):
        if not is_cell_element(child):
            continue

        repeat = repeated_count(child, ATTR_COL_REPEAT)
        start = logical_pos
        end = logical_pos + repeat - 1

        if start <= logical_column <= end:
            offset = logical_column - start

            if repeat == 1:
                if child.tag == COVERED_CELL:
                    replacement = ET.Element(TABLE_CELL)
                    style_name = child.get(ATTR_STYLE_NAME)
                    if style_name:
                        replacement.set(ATTR_STYLE_NAME, style_name)
                    row.remove(child)
                    row.insert(xml_index, replacement)
                    return replacement
                return child

            before_count = offset
            after_count = repeat - offset - 1

            original_style = child.get(ATTR_STYLE_NAME)
            middle = clone_without_repeat(child)
            if middle.tag == COVERED_CELL:
                middle = ET.Element(TABLE_CELL)
                if original_style:
                    middle.set(ATTR_STYLE_NAME, original_style)

            replacement_parts: List[ET.Element] = []

            if before_count:
                before = clone_without_repeat(child)
                if before_count > 1:
                    before.set(ATTR_COL_REPEAT, str(before_count))
                replacement_parts.append(before)

            replacement_parts.append(middle)

            if after_count:
                after = clone_without_repeat(child)
                if after_count > 1:
                    after.set(ATTR_COL_REPEAT, str(after_count))
                replacement_parts.append(after)

            row.remove(child)
            for part_index, part in enumerate(replacement_parts):
                row.insert(xml_index + part_index, part)

            return middle

        logical_pos += repeat

    # Die Zeile ist kürzer als die Zielspalte: leere Zellen ergänzen.
    missing = logical_column - logical_pos
    if missing > 0:
        empty = ET.Element(TABLE_CELL)
        if missing > 1:
            empty.set(ATTR_COL_REPEAT, str(missing))
        row.append(empty)

    target = ET.Element(TABLE_CELL)

    # Stil möglichst von der vorherigen Melderzelle übernehmen.
    previous = get_cell_at(row, logical_column - 1)
    if previous is not None and previous.get(ATTR_STYLE_NAME):
        target.set(ATTR_STYLE_NAME, previous.get(ATTR_STYLE_NAME))

    row.append(target)
    return target


def expanded_rows(table: ET.Element) -> List[ET.Element]:
    """
    Expandiert wiederholte Tabellenzeilen im XML-Baum.
    In normalen Prüflisten sind meist keine wiederholten Datenzeilen vorhanden,
    die Behandlung macht den Parser aber robuster.
    """
    result: List[ET.Element] = []

    for row in list(table):
        if row.tag != TABLE_ROW:
            continue

        repeat = repeated_count(row, ATTR_ROW_REPEAT)

        if repeat == 1:
            result.append(row)
            continue

        # LibreOffice speichert am Tabellenende oft eine leere Zeile mit bis zu
        # 1.048.576 Wiederholungen. Diese darf nicht vollständig aufgefächert
        # werden, da sonst sehr viel Arbeitsspeicher benötigt wird.
        has_visible_content = any(
            cell_text(child).strip()
            for child in row
            if is_cell_element(child)
        )
        if repeat > 1000 and not has_visible_content:
            result.append(row)
            continue

        index = list(table).index(row)
        table.remove(row)

        clones: List[ET.Element] = []
        for _ in range(repeat):
            clone = clone_without_repeat(row)
            clones.append(clone)

        for offset, clone in enumerate(clones):
            table.insert(index + offset, clone)
            result.append(clone)

    return result


# ---------------------- Hekatron-Ereignisspeicher ----------------------

def _xlsx_spaltenname(cell_ref: str) -> str:
    """Gibt aus einer Excel-Zelladresse wie K551 den Spaltennamen K zurück."""
    match = re.match(r"([A-Za-z]+)", cell_ref or "")
    return match.group(1).upper() if match else ""


def _xlsx_zellwert(
    cell: ET.Element,
    shared_strings: List[str],
    ns: Dict[str, str],
) -> str:
    """Liest einen Zellwert aus dem XML einer XLSX-Datei."""
    cell_type = cell.get("t", "")

    if cell_type == "inlineStr":
        return "".join(
            text.text or ""
            for text in cell.findall(".//x:t", ns)
        ).strip()

    value_node = cell.find("x:v", ns)
    if value_node is None or value_node.text is None:
        return ""

    raw = value_node.text.strip()
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return ""

    return raw


def _xlsx_erstes_arbeitsblatt(archive: zipfile.ZipFile) -> str:
    """Ermittelt den XML-Pfad des ersten Arbeitsblatts einer XLSX-Datei."""
    main_ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rel_ns = {"r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    package_rel_ns = {"p": "http://schemas.openxmlformats.org/package/2006/relationships"}

    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    first_sheet = workbook.find(".//x:sheets/x:sheet", main_ns)
    if first_sheet is None:
        raise Auswertungsfehler("Die Excel-Datei enthält kein Arbeitsblatt.")

    relation_id = first_sheet.get(f"{{{rel_ns['r']}}}id")
    if not relation_id:
        raise Auswertungsfehler("Das erste Excel-Arbeitsblatt konnte nicht geöffnet werden.")

    relations = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for relation in relations.findall("p:Relationship", package_rel_ns):
        if relation.get("Id") != relation_id:
            continue
        target = relation.get("Target", "")
        if target.startswith("/"):
            return target.lstrip("/")
        if target.startswith("xl/"):
            return target
        return f"xl/{target.lstrip('/')}"

    raise Auswertungsfehler("Das erste Excel-Arbeitsblatt konnte nicht gefunden werden.")


def _datum_lesen(value: object) -> Optional[datetime]:
    """Liest das Datum des Ereignisspeichers in den üblichen Exportformaten."""
    if isinstance(value, datetime):
        return value

    text = str(value or "").strip()
    if not text:
        return None

    text = re.sub(r"\s+", " ", text)
    formats = (
        "%Y-%m-%d %H:%M:%S",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d",
        "%d.%m.%Y",
    )
    for date_format in formats:
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    return None


def revisionsmelder_aus_xlsx_lesen(
    event_path: str | Path,
    zeitraum_tage: int,
) -> Tuple[Set[Tuple[int, int]], datetime, datetime]:
    """
    Liest den Hekatron-Excel-Export des Ereignisspeichers.

    Auswertung:
    - Spalte „Nr“ = Meldegruppe
    - Spalte „SubNr“ = Meldernummer innerhalb der Gruppe
    - „Status / Befehl“ muss „Revisionsalarm / Rauchalarm“ sein
    - Nur Ereignisse innerhalb des gewählten Zeitraums werden berücksichtigt.

    Der Zeitraum endet beim neuesten Datum im Ereignisspeicher. Bei „1 Tag“
    wird nur der Kalendertag des neuesten Ereignisses ausgewertet, bei „3 Tage“
    dieser Tag und die zwei vorherigen Kalendertage usw. Auch 30 Tage sind möglich.
    """
    if zeitraum_tage not in (1, 3, 7, 30):
        raise Auswertungsfehler("Der Zeitraum muss 1, 3, 7 oder 30 Tage betragen.")

    path = Path(event_path)
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    try:
        with zipfile.ZipFile(path, "r") as archive:
            required = {"xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
            if not required.issubset(set(archive.namelist())):
                raise Auswertungsfehler("Die Datei ist keine gültige Excel-Arbeitsmappe.")

            shared_strings: List[str] = []
            if "xl/sharedStrings.xml" in archive.namelist():
                shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                for item in shared_root.findall("x:si", ns):
                    shared_strings.append(
                        "".join(node.text or "" for node in item.findall(".//x:t", ns))
                    )

            sheet_path = _xlsx_erstes_arbeitsblatt(archive)
            sheet_root = ET.fromstring(archive.read(sheet_path))

            rows: List[Dict[str, str]] = []
            for row in sheet_root.findall(".//x:sheetData/x:row", ns):
                values: Dict[str, str] = {}
                for cell in row.findall("x:c", ns):
                    column = _xlsx_spaltenname(cell.get("r", ""))
                    if column:
                        values[column] = _xlsx_zellwert(cell, shared_strings, ns)
                rows.append(values)

    except zipfile.BadZipFile as exc:
        raise Auswertungsfehler("Die Datei ist keine lesbare XLSX-Datei.") from exc
    except KeyError as exc:
        raise Auswertungsfehler("Die Excel-Datei ist unvollständig oder beschädigt.") from exc
    except ET.ParseError as exc:
        raise Auswertungsfehler("Die Excel-Datei enthält ungültige XML-Daten.") from exc

    if not rows:
        raise Auswertungsfehler("Der Excel-Ereignisspeicher enthält keine Daten.")

    header_columns: Dict[str, str] = {}
    header_index: Optional[int] = None
    for index, row in enumerate(rows[:30]):
        normalized = {value.strip().casefold(): column for column, value in row.items()}
        required_headers = {"nr", "subnr", "status / befehl", "datum"}
        if required_headers.issubset(normalized):
            header_columns = normalized
            header_index = index
            break

    if header_index is None:
        raise Auswertungsfehler(
            "Im Excel-Ereignisspeicher fehlen die Spalten Nr, SubNr, "
            "Status / Befehl oder Datum."
        )

    nr_column = header_columns["nr"]
    subnr_column = header_columns["subnr"]
    status_column = header_columns["status / befehl"]
    date_column = header_columns["datum"]
    type_column = header_columns.get("typ")

    dated_rows: List[Tuple[Dict[str, str], datetime]] = []
    for row in rows[header_index + 1:]:
        event_date = _datum_lesen(row.get(date_column))
        if event_date is not None:
            dated_rows.append((row, event_date))

    if not dated_rows:
        raise Auswertungsfehler(
            "In der Spalte Datum wurden keine lesbaren Datumswerte gefunden."
        )

    newest_event = max(event_date for _, event_date in dated_rows)
    end_date = datetime.combine(newest_event.date(), datetime.max.time())
    start_date = datetime.combine(
        newest_event.date() - timedelta(days=zeitraum_tage - 1),
        datetime.min.time(),
    )

    found: Set[Tuple[int, int]] = set()
    target_status = "revisionsalarm / rauchalarm"

    for row, event_date in dated_rows:
        if event_date < start_date or event_date > end_date:
            continue

        status = row.get(status_column, "").strip().casefold()
        if status != target_status:
            continue

        if type_column and row.get(type_column, "").strip().casefold() != "meldergruppe":
            continue

        group = as_int(row.get(nr_column))
        detector = as_int(row.get(subnr_column))
        if group is None or detector is None:
            continue
        if group <= 0 or detector <= 0:
            continue

        found.add((group, detector))

    return found, start_date, end_date


def feuermelder_aus_esser_pdf_lesen(
    event_path: str | Path,
    zeitraum_tage: int,
    voralarm_beruecksichtigen: bool = False,
) -> Tuple[Set[Tuple[int, int]], datetime, datetime]:
    """Liest einen Esser-Ereignisspeicher im PDF-Format.

    Standardmäßig werden ausschließlich Zeilen mit dem exakten Ereignis
    ``FEUER`` berücksichtigt. Ist ``voralarm_beruecksichtigen`` aktiviert,
    zählen zusätzlich ``VORALARM`` und ``VORALARM-ENDE`` – jedoch nur Zeilen,
    in denen sowohl Gruppe als auch Meldernummer vorhanden sind.
    ``FEUER-ENDE`` wird nicht gewertet.
    """
    if zeitraum_tage not in (1, 3, 7, 30):
        raise Auswertungsfehler("Der Zeitraum muss 1, 3, 7 oder 30 Tage betragen.")
    if PdfReader is None:
        raise Auswertungsfehler(
            "Für Esser-PDF-Dateien fehlt das Modul pypdf.\n\n"
            "Bitte unter Windows einmal ausführen:\npython -m pip install pypdf"
        )

    path = Path(event_path)
    try:
        reader = PdfReader(str(path))
        page_texts = [(page.extract_text() or "") for page in reader.pages]
    except Exception as exc:
        raise Auswertungsfehler(f"Die Esser-PDF-Datei konnte nicht gelesen werden: {exc}") from exc

    full_text = "\n".join(page_texts)
    # Das Exportjahr steht im Kopf-/Fußbereich, z. B. 30.07.2026.
    years = [int(y) for y in re.findall(r"\b\d{1,2}\.\d{1,2}\.(20\d{2})\b", full_text)]
    year = max(years) if years else datetime.now().year

    all_dates: List[datetime] = []
    found_rows: List[Tuple[datetime, int, int]] = []
    erlaubte_ereignisse = ["FEUER"]
    if voralarm_beruecksichtigen:
        erlaubte_ereignisse.extend(["VORALARM", "VORALARM-ENDE"])
    ereignis_muster = "(?:" + "|".join(re.escape(e) for e in erlaubte_ereignisse) + ")"

    row_pattern = re.compile(
        r"^\s*\d+?(\d{2}\.\d{2}\.)"
        r"(\d{1,2}:\d{2}:\d{2})\s*" + ereignis_muster + r"\s+"
        r"Gruppe:\s*(\d+)\s*Melder:\s*(\d+)\s*$",
        re.IGNORECASE,
    )
    date_pattern = re.compile(
        r"^\s*\d+?(\d{2}\.\d{2}\.)"
        r"(\d{1,2}:\d{2}:\d{2})"
    )

    for text in page_texts:
        for raw_line in text.splitlines():
            line = re.sub(r"\s+", " ", raw_line).strip()
            dm = date_pattern.match(line)
            if dm:
                try:
                    all_dates.append(datetime.strptime(f"{dm.group(1)}{year} {dm.group(2)}", "%d.%m.%Y %H:%M:%S"))
                except ValueError:
                    pass
            match = row_pattern.match(line)
            if not match:
                continue
            try:
                event_date = datetime.strptime(
                    f"{match.group(1)}{year} {match.group(2)}",
                    "%d.%m.%Y %H:%M:%S",
                )
                group = int(match.group(3))
                detector = int(match.group(4))
            except ValueError:
                continue
            if group > 0 and detector > 0:
                found_rows.append((event_date, group, detector))

    if not all_dates:
        raise Auswertungsfehler("Im Esser-PDF wurden keine lesbaren Ereignisdatumswerte gefunden.")

    newest_event = max(all_dates)
    end_date = datetime.combine(newest_event.date(), datetime.max.time())
    start_date = datetime.combine(
        newest_event.date() - timedelta(days=zeitraum_tage - 1),
        datetime.min.time(),
    )
    found = {
        (group, detector)
        for event_date, group, detector in found_rows
        if start_date <= event_date <= end_date
    }
    return found, start_date, end_date


def revisionsmelder_lesen(
    event_path: str | Path,
    zeitraum_tage: int,
    manufacturer: str = "Hekatron",
    esser_voralarm: bool = False,
) -> Tuple[Set[Tuple[int, int]], datetime, datetime]:
    """Wählt den Ereignisspeicher-Leser passend zum Hersteller aus."""
    path = Path(event_path)
    if not path.is_file():
        raise Auswertungsfehler("Der Ereignisspeicher wurde nicht gefunden.")

    if manufacturer.casefold() == "esser":
        if path.suffix.lower() == ".pdf":
            return feuermelder_aus_esser_pdf_lesen(path, zeitraum_tage, esser_voralarm)
        raise Auswertungsfehler("Bitte für Esser einen Ereignisspeicher im PDF-Format auswählen.")

    if path.suffix.lower() == ".xlsx":
        return revisionsmelder_aus_xlsx_lesen(path, zeitraum_tage)

    raise Auswertungsfehler(
        "Bitte für Hekatron den Ereignisspeicher als Excel-Datei (.xlsx) auswählen."
    )


def txt_pruefbericht_schreiben(result: Dict[str, object], quarter: int) -> str:
    """Schreibt eine übersichtliche TXT-Datei neben die ausgefüllte Prüfliste."""
    output_path = Path(str(result["output"]))
    report_path = output_path.with_name(output_path.stem + "_Pruefbericht.txt")

    unchecked = list(result.get("unchecked", []))
    incomplete = dict(result.get("incomplete_groups", {}))
    missing_groups = list(result.get("missing_groups", []))

    lines: List[str] = [
        "BMA Wartungsauswertung - Prüfbericht",
        f"Version: {VERSION}",
        f"Quartal: {quarter}",
        f"Auswertungszeitraum: {result.get('zeitraum_tage', '')} Tag(e)",
        f"Von: {result.get('start_date'):%d.%m.%Y}" if result.get("start_date") else "",
        f"Bis: {result.get('end_date'):%d.%m.%Y}" if result.get("end_date") else "",
        "",
        f"Im Ereignisspeicher gefunden: {result.get('found', 0)}",
        f"Neu eingetragen: {len(result.get('inserted', []))}",
        f"Bereits belegt: {len(result.get('occupied', []))}",
        f"Nicht geprüft: {len(unchecked)}",
        f"Unvollständige Meldegruppen: {len(incomplete)}",
        "",
        "NICHT GEPRÜFTE MELDER",
        "----------------------",
    ]

    if unchecked:
        lines.extend(f"Meldegruppe {group}, Melder {detector}" for group, detector in unchecked)
    else:
        lines.append("Keine - alle in der Prüfliste vorhandenen Melder sind geprüft.")

    lines.extend(["", "UNVOLLSTÄNDIGE MELDEGRUPPEN", "---------------------------"])
    if incomplete:
        for group in sorted(incomplete):
            detectors = incomplete[group]
            lines.append(
                f"Meldegruppe {group}: Es fehlen Melder "
                + ", ".join(str(number) for number in detectors)
            )
    else:
        lines.append("Keine - alle Meldegruppen sind vollständig geprüft.")

    lines.extend(["", "MELDER OHNE PASSENDE MELDEGRUPPE IN DER PRÜFLISTE", "-------------------------------------------------"])
    if missing_groups:
        lines.extend(f"Meldegruppe {group}, Melder {detector}" for group, detector in missing_groups)
    else:
        lines.append("Keine.")

    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    return str(report_path)


# ----------------------------- ODS-Bearbeitung -----------------------------

def find_first_sheet(root: ET.Element) -> ET.Element:
    spreadsheet = root.find(
        f".//{{{NS['office']}}}spreadsheet"
    )
    if spreadsheet is None:
        raise Auswertungsfehler("Die ODS-Datei enthält keine Tabelle.")

    table = spreadsheet.find(TABLE_ROW.replace("table-row", "table"))
    if table is None:
        # Direkte Suche ist stabiler bei unterschiedlichen XML-Strukturen.
        table = spreadsheet.find(f"{{{NS['table']}}}table")

    if table is None:
        raise Auswertungsfehler("Die ODS-Datei enthält kein Arbeitsblatt.")

    return table



def find_checklist_sheet(root: ET.Element, checklist_version: str) -> ET.Element:
    """Wählt das Arbeitsblatt passend zur Prüflistenversion aus."""
    if checklist_version == "Version 1":
        return find_first_sheet(root)

    spreadsheet = root.find(f".//{{{NS['office']}}}spreadsheet")
    if spreadsheet is None:
        raise Auswertungsfehler("Die ODS-Datei enthält keine Tabelle.")

    name_attr = f"{{{NS['table']}}}name"
    tables = spreadsheet.findall(f"{{{NS['table']}}}table")
    for table in tables:
        sheet_name = (table.get(name_attr) or "").strip().lower()
        if sheet_name in {"prüfliste", "pruefliste"}:
            return table

    raise Auswertungsfehler(
        "Für Prüflisten-Version 2 wurde kein Arbeitsblatt mit dem Namen "
        "'Prüfliste' gefunden."
    )

def ods_pruefliste_ausfuellen(
    event_path: str | Path,
    ods_path: str | Path,
    output_path: str | Path,
    quarter: int,
    zeitraum_tage: int,
    manufacturer: str = "Hekatron",
    esser_voralarm: bool = False,
    checklist_version: str = "Version 1",
) -> Dict[str, object]:
    if quarter not in (1, 2, 3, 4):
        raise Auswertungsfehler("Das Quartal muss 1, 2, 3 oder 4 sein.")

    event_detectors, start_date, end_date = revisionsmelder_lesen(event_path, zeitraum_tage, manufacturer, esser_voralarm)

    ods_path = Path(ods_path)
    output_path = Path(output_path)

    if not ods_path.is_file():
        raise Auswertungsfehler("Die ODS-Prüfliste wurde nicht gefunden.")

    if ods_path.suffix.lower() != ".ods":
        raise Auswertungsfehler("Bitte eine Prüfliste im Format .ods auswählen.")

    try:
        with zipfile.ZipFile(ods_path, "r") as source:
            if "content.xml" not in source.namelist():
                raise Auswertungsfehler(
                    "Die Datei ist keine gültige ODS-Arbeitsmappe."
                )

            content_root = ET.fromstring(source.read("content.xml"))
            table = find_checklist_sheet(content_root, checklist_version)
            rows = expanded_rows(table)

            # Meldegruppe -> (Zeile, Anzahl Melder)
            groups: Dict[int, Tuple[ET.Element, Optional[int]]] = {}

            for row in rows:
                group_cell = get_cell_at(row, MG_COLUMN_INDEX)
                if group_cell is None:
                    continue

                group = as_int(cell_text(group_cell))
                if group is None:
                    continue

                count_cell = get_cell_at(row, COUNT_COLUMN_INDEX)
                count = as_int(cell_text(count_cell)) if count_cell is not None else None

                groups[group] = (row, count)

            if not groups:
                raise Auswertungsfehler(
                    f"In Spalte A des gewählten Prüflistenblatts ({checklist_version}) wurden keine Meldegruppen gefunden."
                )

            inserted: List[Tuple[int, int]] = []
            occupied: List[Tuple[int, int, str]] = []
            missing_groups: List[Tuple[int, int]] = []
            for group, detector in sorted(event_detectors):
                group_info = groups.get(group)

                if group_info is None:
                    missing_groups.append((group, detector))
                    continue

                row, maximum = group_info

                # Die Spalte "Anzahl Melder" wird bewusst nicht als harte Grenze verwendet.
                # Manche Vorlagen enthalten dort unvollständige oder anders berechnete Werte.
                # Jeder im Ereignisspeicher gefundene Melder wird daher eingetragen.
                target_column = FIRST_DETECTOR_COLUMN_INDEX + detector - 1
                target_cell = ensure_cell_at(row, target_column)
                previous_value = cell_text(target_cell).strip()

                if previous_value:
                    occupied.append((group, detector, previous_value))
                    continue

                set_numeric_cell_value(target_cell, quarter)
                inserted.append((group, detector))

            unchecked: List[Tuple[int, int]] = []
            incomplete_groups: Dict[int, List[int]] = {}
            for group, (row, maximum) in sorted(groups.items()):
                if maximum is None or maximum <= 0:
                    continue
                missing_detectors: List[int] = []
                for detector in range(1, maximum + 1):
                    detector_cell = get_cell_at(
                        row, FIRST_DETECTOR_COLUMN_INDEX + detector - 1
                    )
                    value = cell_text(detector_cell).strip() if detector_cell is not None else ""
                    if not value:
                        unchecked.append((group, detector))
                        missing_detectors.append(detector)
                if missing_detectors:
                    incomplete_groups[group] = missing_detectors

            content_bytes = ET.tostring(
                content_root,
                encoding="utf-8",
                xml_declaration=True,
            )

            output_path.parent.mkdir(parents=True, exist_ok=True)
            temporary = output_path.with_suffix(".ods.tmp")

            with zipfile.ZipFile(temporary, "w") as target:
                # ODS-Vorgabe: mimetype zuerst und unkomprimiert.
                if "mimetype" in source.namelist():
                    mimetype_data = source.read("mimetype")
                    target.writestr(
                        "mimetype",
                        mimetype_data,
                        compress_type=zipfile.ZIP_STORED,
                    )

                for item in source.infolist():
                    if item.filename == "mimetype":
                        continue

                    data = source.read(item.filename)
                    if item.filename == "content.xml":
                        data = content_bytes

                    target.writestr(
                        item,
                        data,
                        compress_type=zipfile.ZIP_DEFLATED,
                    )

            os.replace(temporary, output_path)

    except zipfile.BadZipFile as exc:
        raise Auswertungsfehler(
            "Die gewählte Datei ist keine gültige ODS-Datei."
        ) from exc
    except PermissionError as exc:
        raise Auswertungsfehler(
            "Die Ausgabedatei kann nicht gespeichert werden. "
            "Bitte schließe sie in LibreOffice und versuche es erneut."
        ) from exc

    return {
        "found": len(event_detectors),
        "inserted": inserted,
        "occupied": occupied,
        "missing_groups": missing_groups,
        "unchecked": unchecked,
        "incomplete_groups": incomplete_groups,
        "output": str(output_path),
        "zeitraum_tage": zeitraum_tage,
        "start_date": start_date,
        "end_date": end_date,
    }



def _xlsx_spaltenindex(column_name: str) -> int:
    """Wandelt einen Excel-Spaltennamen wie F oder AA in eine 1-basierte Nummer um."""
    value = 0
    for character in column_name.upper():
        if not ("A" <= character <= "Z"):
            continue
        value = value * 26 + (ord(character) - ord("A") + 1)
    return value


def _xlsx_spaltenbuchstaben(column_index: int) -> str:
    """Wandelt eine 1-basierte Spaltennummer in einen Excel-Spaltennamen um."""
    if column_index < 1:
        raise ValueError("Die Spaltennummer muss mindestens 1 sein.")
    result = ""
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _xlsx_zelle_in_zeile(row: ET.Element, column_index: int) -> Optional[ET.Element]:
    for cell in row.findall("x:c", XLSX_NS):
        if _xlsx_spaltenindex(_xlsx_spaltenname(cell.get("r", ""))) == column_index:
            return cell
    return None


def _xlsx_zelle_sicherstellen(row: ET.Element, row_number: int, column_index: int) -> ET.Element:
    """Erzeugt eine Excel-Zelle an der richtigen Position und übernimmt möglichst den Stil."""
    existing = _xlsx_zelle_in_zeile(row, column_index)
    if existing is not None:
        return existing

    reference = f"{_xlsx_spaltenbuchstaben(column_index)}{row_number}"
    new_cell = ET.Element(f"{{{XLSX_NS['x']}}}c", {"r": reference})

    # Format möglichst von der vorherigen Melderzelle übernehmen.
    style_source = _xlsx_zelle_in_zeile(row, column_index - 1)
    if style_source is None:
        style_source = _xlsx_zelle_in_zeile(row, FIRST_DETECTOR_COLUMN_INDEX)
    if style_source is not None and style_source.get("s") is not None:
        new_cell.set("s", style_source.get("s"))

    cells = list(row.findall("x:c", XLSX_NS))
    inserted = False
    for index, cell in enumerate(cells):
        current_index = _xlsx_spaltenindex(_xlsx_spaltenname(cell.get("r", "")))
        if current_index > column_index:
            row.insert(list(row).index(cell), new_cell)
            inserted = True
            break
    if not inserted:
        row.append(new_cell)
    return new_cell


def _xlsx_numerischen_wert_setzen(cell: ET.Element, number: int) -> None:
    """Setzt einen numerischen Wert, ohne den vorhandenen Zellstil zu löschen."""
    cell.attrib.pop("t", None)
    for child in list(cell):
        if child.tag in {
            f"{{{XLSX_NS['x']}}}v",
            f"{{{XLSX_NS['x']}}}f",
            f"{{{XLSX_NS['x']}}}is",
        }:
            cell.remove(child)
    value = ET.SubElement(cell, f"{{{XLSX_NS['x']}}}v")
    value.text = str(number)


XLSX_NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
ET.register_namespace("", XLSX_NS["x"])


def xlsx_pruefliste_ausfuellen(
    event_path: str | Path,
    xlsx_path: str | Path,
    output_path: str | Path,
    quarter: int,
    zeitraum_tage: int,
    manufacturer: str = "Hekatron",
    esser_voralarm: bool = False,
    checklist_version: str = "Version 1",
) -> Dict[str, object]:
    """Füllt eine Excel-Prüfliste mit openpyxl Windows-/Excel-kompatibel aus."""
    if quarter not in (1, 2, 3, 4):
        raise Auswertungsfehler("Das Quartal muss 1, 2, 3 oder 4 sein.")

    if load_workbook is None:
        raise Auswertungsfehler(
            "Für Excel-Prüflisten fehlt das Modul openpyxl.\n\n"
            "Bitte unter Windows einmal ausführen:\n"
            "python -m pip install openpyxl"
        )

    event_detectors, start_date, end_date = revisionsmelder_lesen(event_path, zeitraum_tage, manufacturer, esser_voralarm)
    source_path = Path(xlsx_path)
    output_path = Path(output_path)

    if not source_path.is_file():
        raise Auswertungsfehler("Die Excel-Prüfliste wurde nicht gefunden.")
    if source_path.suffix.lower() != ".xlsx":
        raise Auswertungsfehler("Bitte eine Excel-Prüfliste im Format .xlsx auswählen.")
    if output_path.suffix.lower() != ".xlsx":
        raise Auswertungsfehler("Die fertige Excel-Prüfliste muss als .xlsx gespeichert werden.")

    temporary = output_path.with_name(output_path.stem + "_tmp.xlsx")

    try:
        # data_only=False erhält Formeln. keep_links=True bewahrt externe Verknüpfungen.
        workbook = load_workbook(
            filename=source_path,
            read_only=False,
            data_only=False,
            keep_links=True,
        )

        if not workbook.worksheets:
            raise Auswertungsfehler("Die Excel-Prüfliste enthält kein Arbeitsblatt.")

        if checklist_version == "Version 2":
            worksheet = next(
                (ws for ws in workbook.worksheets if ws.title.strip().lower() in {"prüfliste", "pruefliste"}),
                None,
            )
            if worksheet is None:
                raise Auswertungsfehler(
                    "Für Prüflisten-Version 2 wurde kein Arbeitsblatt mit dem Namen 'Prüfliste' gefunden."
                )
        else:
            worksheet = workbook.worksheets[0]

        groups: Dict[int, Tuple[int, Optional[int]]] = {}
        for row_number in range(1, worksheet.max_row + 1):
            group = as_int(worksheet.cell(row=row_number, column=MG_COLUMN_INDEX).value)
            if group is not None:
                maximum = as_int(worksheet.cell(row=row_number, column=COUNT_COLUMN_INDEX).value)
                groups[group] = (row_number, maximum)

        if not groups:
            raise Auswertungsfehler(f"In Spalte A des gewählten Prüflistenblatts ({checklist_version}) wurden keine Meldegruppen gefunden.")

        inserted: List[Tuple[int, int]] = []
        occupied: List[Tuple[int, int, str]] = []
        missing_groups: List[Tuple[int, int]] = []

        for group, detector in sorted(event_detectors):
            group_info = groups.get(group)
            if group_info is None:
                missing_groups.append((group, detector))
                continue

            row_number, _maximum = group_info
            target_column = FIRST_DETECTOR_COLUMN_INDEX + detector - 1
            target_cell = worksheet.cell(row=row_number, column=target_column)
            previous_value = target_cell.value

            if previous_value is not None and str(previous_value).strip() != "":
                occupied.append((group, detector, str(previous_value)))
                continue

            target_cell.value = quarter
            inserted.append((group, detector))

        unchecked: List[Tuple[int, int]] = []
        incomplete_groups: Dict[int, List[int]] = {}
        for group, (row_number, maximum) in sorted(groups.items()):
            if maximum is None or maximum <= 0:
                continue
            missing_detectors: List[int] = []
            for detector in range(1, maximum + 1):
                value = worksheet.cell(
                    row=row_number,
                    column=FIRST_DETECTOR_COLUMN_INDEX + detector - 1,
                ).value
                if value is None or str(value).strip() == "":
                    unchecked.append((group, detector))
                    missing_detectors.append(detector)
            if missing_detectors:
                incomplete_groups[group] = missing_detectors

        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Zuerst unter einem echten .xlsx-Namen speichern und anschließend atomar ersetzen.
        workbook.save(temporary)
        workbook.close()

        # Sofortige Integritätsprüfung: Die erzeugte Datei muss erneut lesbar sein.
        check_workbook = load_workbook(
            filename=temporary,
            read_only=True,
            data_only=False,
            keep_links=True,
        )
        check_workbook.close()

        os.replace(temporary, output_path)

    except PermissionError as exc:
        raise Auswertungsfehler(
            "Die Ausgabedatei kann nicht gespeichert werden. Bitte schließe sie in Excel "
            "und versuche es erneut."
        ) from exc
    except InvalidFileException as exc:
        raise Auswertungsfehler(
            "Die gewählte Prüfliste ist keine gültige Excel-.xlsx-Datei."
        ) from exc
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise Auswertungsfehler(
            "Die Excel-Prüfliste konnte nicht fehlerfrei verarbeitet werden. "
            "Die Originaldatei wurde nicht verändert."
        ) from exc
    finally:
        try:
            if temporary.exists():
                temporary.unlink()
        except OSError:
            pass

    return {
        "found": len(event_detectors),
        "inserted": inserted,
        "occupied": occupied,
        "missing_groups": missing_groups,
        "unchecked": unchecked,
        "incomplete_groups": incomplete_groups,
        "output": str(output_path),
        "zeitraum_tage": zeitraum_tage,
        "start_date": start_date,
        "end_date": end_date,
    }



def _xls_wert_mit_stil_schreiben(worksheet, row_index: int, column_index: int, value: object) -> None:
    """Überschreibt eine XLS-Zelle und übernimmt nach Möglichkeit ihren bisherigen Stil."""
    old_xf_index = None
    try:
        old_row = worksheet._Worksheet__rows.get(row_index)
        if old_row is not None:
            old_cell = old_row._Row__cells.get(column_index)
            if old_cell is not None:
                old_xf_index = old_cell.xf_idx
    except Exception:
        old_xf_index = None

    worksheet.write(row_index, column_index, value)

    if old_xf_index is not None:
        try:
            new_row = worksheet._Worksheet__rows.get(row_index)
            new_cell = new_row._Row__cells.get(column_index)
            new_cell.xf_idx = old_xf_index
        except Exception:
            pass


def xls_pruefliste_ausfuellen(
    event_path: str | Path,
    xls_path: str | Path,
    output_path: str | Path,
    quarter: int,
    zeitraum_tage: int,
    manufacturer: str = "Hekatron",
    esser_voralarm: bool = False,
    checklist_version: str = "Version 1",
) -> Dict[str, object]:
    """Füllt eine klassische Excel-97–2003-Prüfliste (.xls) aus."""
    if quarter not in (1, 2, 3, 4):
        raise Auswertungsfehler("Das Quartal muss 1, 2, 3 oder 4 sein.")

    if xlrd is None or copy_xls_workbook is None:
        raise Auswertungsfehler(
            "Für alte Excel-.xls-Prüflisten fehlen zusätzliche Module.\n\n"
            "Bitte unter Windows einmal ausführen:\n"
            "python -m pip install xlrd xlwt xlutils"
        )

    event_detectors, start_date, end_date = revisionsmelder_lesen(
        event_path, zeitraum_tage, manufacturer, esser_voralarm
    )
    source_path = Path(xls_path)
    output_path = Path(output_path)

    if not source_path.is_file():
        raise Auswertungsfehler("Die Excel-.xls-Prüfliste wurde nicht gefunden.")
    if source_path.suffix.lower() != ".xls":
        raise Auswertungsfehler("Bitte eine Prüfliste im alten Excel-Format .xls auswählen.")
    if output_path.suffix.lower() != ".xls":
        raise Auswertungsfehler("Eine .xls-Vorlage muss wieder als .xls gespeichert werden.")

    temporary = output_path.with_name(output_path.stem + "_tmp.xls")

    try:
        read_book = xlrd.open_workbook(str(source_path), formatting_info=True)
        if read_book.nsheets <= 0:
            raise Auswertungsfehler("Die Excel-.xls-Prüfliste enthält kein Arbeitsblatt.")

        if checklist_version == "Version 2":
            sheet_index = next(
                (
                    index for index, name in enumerate(read_book.sheet_names())
                    if name.strip().lower() in {"prüfliste", "pruefliste"}
                ),
                None,
            )
            if sheet_index is None:
                raise Auswertungsfehler(
                    "Für Prüflisten-Version 2 wurde kein Arbeitsblatt mit dem Namen 'Prüfliste' gefunden."
                )
        else:
            sheet_index = 0

        read_sheet = read_book.sheet_by_index(sheet_index)
        write_book = copy_xls_workbook(read_book)
        write_sheet = write_book.get_sheet(sheet_index)

        groups: Dict[int, Tuple[int, Optional[int]]] = {}
        for row_index in range(read_sheet.nrows):
            group = as_int(read_sheet.cell_value(row_index, MG_COLUMN_INDEX - 1))
            if group is not None:
                maximum = as_int(read_sheet.cell_value(row_index, COUNT_COLUMN_INDEX - 1))
                groups[group] = (row_index, maximum)

        if not groups:
            raise Auswertungsfehler(
                f"In Spalte A des gewählten Prüflistenblatts ({checklist_version}) wurden keine Meldegruppen gefunden."
            )

        inserted: List[Tuple[int, int]] = []
        occupied: List[Tuple[int, int, str]] = []
        missing_groups: List[Tuple[int, int]] = []
        written_values: Dict[Tuple[int, int], object] = {}

        for group, detector in sorted(event_detectors):
            group_info = groups.get(group)
            if group_info is None:
                missing_groups.append((group, detector))
                continue

            row_index, _maximum = group_info
            column_index = FIRST_DETECTOR_COLUMN_INDEX + detector - 2
            previous_value = (
                written_values.get((row_index, column_index))
                if (row_index, column_index) in written_values
                else read_sheet.cell_value(row_index, column_index)
                if column_index < read_sheet.ncols
                else ""
            )

            if previous_value is not None and str(previous_value).strip() != "":
                occupied.append((group, detector, str(previous_value)))
                continue

            _xls_wert_mit_stil_schreiben(write_sheet, row_index, column_index, quarter)
            written_values[(row_index, column_index)] = quarter
            inserted.append((group, detector))

        unchecked: List[Tuple[int, int]] = []
        incomplete_groups: Dict[int, List[int]] = {}
        for group, (row_index, maximum) in sorted(groups.items()):
            if maximum is None or maximum <= 0:
                continue
            missing_detectors: List[int] = []
            for detector in range(1, maximum + 1):
                column_index = FIRST_DETECTOR_COLUMN_INDEX + detector - 2
                if (row_index, column_index) in written_values:
                    value = written_values[(row_index, column_index)]
                elif column_index < read_sheet.ncols:
                    value = read_sheet.cell_value(row_index, column_index)
                else:
                    value = ""
                if value is None or str(value).strip() == "":
                    unchecked.append((group, detector))
                    missing_detectors.append(detector)
            if missing_detectors:
                incomplete_groups[group] = missing_detectors

        output_path.parent.mkdir(parents=True, exist_ok=True)
        write_book.save(str(temporary))

        # Integritätsprüfung: Die erzeugte XLS-Datei muss sich erneut einlesen lassen.
        check_book = xlrd.open_workbook(str(temporary), on_demand=True)
        check_book.release_resources()
        os.replace(temporary, output_path)

    except PermissionError as exc:
        raise Auswertungsfehler(
            "Die Ausgabedatei kann nicht gespeichert werden. Bitte schließe sie in Excel und versuche es erneut."
        ) from exc
    except Auswertungsfehler:
        raise
    except Exception as exc:
        raise Auswertungsfehler(
            "Die Excel-.xls-Prüfliste konnte nicht fehlerfrei verarbeitet werden. "
            "Die Originaldatei wurde nicht verändert."
        ) from exc
    finally:
        try:
            if temporary.exists():
                temporary.unlink()
        except OSError:
            pass

    return {
        "found": len(event_detectors),
        "inserted": inserted,
        "occupied": occupied,
        "missing_groups": missing_groups,
        "unchecked": unchecked,
        "incomplete_groups": incomplete_groups,
        "output": str(output_path),
        "zeitraum_tage": zeitraum_tage,
        "start_date": start_date,
        "end_date": end_date,
    }


def pruefliste_ausfuellen(
    event_path: str | Path,
    checklist_path: str | Path,
    output_path: str | Path,
    quarter: int,
    zeitraum_tage: int,
    manufacturer: str = "Hekatron",
    esser_voralarm: bool = False,
    checklist_version: str = "Version 1",
) -> Dict[str, object]:
    """Wählt den passenden Prüflisten-Schreiber anhand der Dateiendung."""
    suffix = Path(checklist_path).suffix.lower()
    output_suffix = Path(output_path).suffix.lower()

    if suffix == ".ods":
        if output_suffix != ".ods":
            raise Auswertungsfehler("Eine ODS-Vorlage muss wieder als .ods gespeichert werden.")
        return ods_pruefliste_ausfuellen(event_path, checklist_path, output_path, quarter, zeitraum_tage, manufacturer, esser_voralarm, checklist_version)

    if suffix == ".xlsx":
        if output_suffix != ".xlsx":
            raise Auswertungsfehler("Eine Excel-Vorlage muss wieder als .xlsx gespeichert werden.")
        return xlsx_pruefliste_ausfuellen(event_path, checklist_path, output_path, quarter, zeitraum_tage, manufacturer, esser_voralarm, checklist_version)

    if suffix == ".xls":
        if output_suffix != ".xls":
            raise Auswertungsfehler("Eine Excel-.xls-Vorlage muss wieder als .xls gespeichert werden.")
        return xls_pruefliste_ausfuellen(event_path, checklist_path, output_path, quarter, zeitraum_tage, manufacturer, esser_voralarm, checklist_version)

    raise Auswertungsfehler("Bitte eine Prüfliste im Format .ods, .xlsx oder .xls auswählen.")



def prueflisten_statistik(checklist_path: str | Path, checklist_version: str = "Version 1") -> Dict[str, int]:
    """Ermittelt Q1-Q4, Gesamtzahl und offene Melder aus einer Prüfliste."""
    path = Path(checklist_path)
    if not path.is_file():
        raise Auswertungsfehler("Die Prüfliste wurde nicht gefunden.")

    stats = {"q1": 0, "q2": 0, "q3": 0, "q4": 0, "total": 0, "open": 0}

    if path.suffix.lower() == ".xlsx":
        if load_workbook is None:
            raise Auswertungsfehler("Für Excel-Prüflisten fehlt das Modul openpyxl.")
        try:
            workbook = load_workbook(path, read_only=True, data_only=True, keep_links=True)
            if checklist_version == "Version 2":
                worksheet = next(
                    (ws for ws in workbook.worksheets if ws.title.strip().lower() in {"prüfliste", "pruefliste"}),
                    None,
                )
                if worksheet is None:
                    raise Auswertungsfehler(
                        "Für Prüflisten-Version 2 wurde kein Arbeitsblatt mit dem Namen 'Prüfliste' gefunden."
                    )
            else:
                worksheet = workbook.worksheets[0]
            for row_number in range(1, worksheet.max_row + 1):
                group = as_int(worksheet.cell(row=row_number, column=MG_COLUMN_INDEX).value)
                maximum = as_int(worksheet.cell(row=row_number, column=COUNT_COLUMN_INDEX).value)
                if group is None or maximum is None or maximum <= 0:
                    continue
                stats["total"] += maximum
                for detector in range(1, maximum + 1):
                    value = worksheet.cell(
                        row=row_number,
                        column=FIRST_DETECTOR_COLUMN_INDEX + detector - 1,
                    ).value
                    text = str(value).strip() if value is not None else ""
                    if text in {"1", "2", "3", "4"}:
                        stats[f"q{text}"] += 1
                    elif not text:
                        stats["open"] += 1
            workbook.close()
        except Exception as exc:
            raise Auswertungsfehler("Die Excel-Prüfliste konnte für die Statistik nicht gelesen werden.") from exc
        return stats

    if path.suffix.lower() == ".xls":
        if xlrd is None:
            raise Auswertungsfehler(
                "Für alte Excel-.xls-Prüflisten fehlt das Modul xlrd.\n\n"
                "Bitte ausführen: python -m pip install xlrd xlwt xlutils"
            )
        try:
            workbook = xlrd.open_workbook(str(path), on_demand=True)
            if checklist_version == "Version 2":
                sheet_index = next(
                    (
                        index for index, name in enumerate(workbook.sheet_names())
                        if name.strip().lower() in {"prüfliste", "pruefliste"}
                    ),
                    None,
                )
                if sheet_index is None:
                    raise Auswertungsfehler(
                        "Für Prüflisten-Version 2 wurde kein Arbeitsblatt mit dem Namen 'Prüfliste' gefunden."
                    )
            else:
                sheet_index = 0
            worksheet = workbook.sheet_by_index(sheet_index)
            for row_index in range(worksheet.nrows):
                group = as_int(worksheet.cell_value(row_index, MG_COLUMN_INDEX - 1))
                maximum = as_int(worksheet.cell_value(row_index, COUNT_COLUMN_INDEX - 1))
                if group is None or maximum is None or maximum <= 0:
                    continue
                stats["total"] += maximum
                for detector in range(1, maximum + 1):
                    column_index = FIRST_DETECTOR_COLUMN_INDEX + detector - 2
                    value = worksheet.cell_value(row_index, column_index) if column_index < worksheet.ncols else ""
                    text = str(value).strip()
                    if text.endswith(".0") and text[:-2] in {"1", "2", "3", "4"}:
                        text = text[:-2]
                    if text in {"1", "2", "3", "4"}:
                        stats[f"q{text}"] += 1
                    elif not text:
                        stats["open"] += 1
            workbook.release_resources()
        except Auswertungsfehler:
            raise
        except Exception as exc:
            raise Auswertungsfehler("Die Excel-.xls-Prüfliste konnte für die Statistik nicht gelesen werden.") from exc
        return stats

    if path.suffix.lower() == ".ods":
        try:
            with zipfile.ZipFile(path, "r") as archive:
                root = ET.fromstring(archive.read("content.xml"))
                table = find_checklist_sheet(root, checklist_version)
                for row in expanded_rows(table):
                    group_cell = get_cell_at(row, MG_COLUMN_INDEX)
                    count_cell = get_cell_at(row, COUNT_COLUMN_INDEX)
                    group = as_int(cell_text(group_cell)) if group_cell is not None else None
                    maximum = as_int(cell_text(count_cell)) if count_cell is not None else None
                    if group is None or maximum is None or maximum <= 0:
                        continue
                    stats["total"] += maximum
                    for detector in range(1, maximum + 1):
                        detector_cell = get_cell_at(
                            row, FIRST_DETECTOR_COLUMN_INDEX + detector - 1
                        )
                        text = cell_text(detector_cell).strip() if detector_cell is not None else ""
                        if text in {"1", "2", "3", "4"}:
                            stats[f"q{text}"] += 1
                        elif not text:
                            stats["open"] += 1
        except Exception as exc:
            raise Auswertungsfehler("Die ODS-Prüfliste konnte für die Statistik nicht gelesen werden.") from exc
        return stats

    raise Auswertungsfehler("Bitte eine Prüfliste im Format .ods, .xlsx oder .xls auswählen.")


def result_text(result: Dict[str, object]) -> str:
    inserted = result["inserted"]
    occupied = result["occupied"]
    missing = result["missing_groups"]
    unchecked = result.get("unchecked", [])
    incomplete = result.get("incomplete_groups", {})
    lines = [
        f"Auswertungszeitraum: {result.get('start_date'):%d.%m.%Y} bis {result.get('end_date'):%d.%m.%Y}",
        f"Revisionsmelder gefunden: {result['found']}",
        f"Neu eingetragen: {len(inserted)}",
        f"Bereits belegt und nicht überschrieben: {len(occupied)}",
        f"Meldegruppe nicht in der Prüfliste: {len(missing)}",
        f"Nicht geprüfte Melder: {len(unchecked)}",
        f"Unvollständige Meldegruppen: {len(incomplete)}",
        "",
        f"Prüfliste gespeichert unter:\n{result['output']}",
    ]

    if result.get("report"):
        lines.append(f"Prüfbericht gespeichert unter:\n{result['report']}")
    else:
        lines.append("TXT-Prüfbericht: nicht erstellt")

    if missing:
        lines.extend([
            "",
            "Fehlende Gruppen: "
            + ", ".join(f"{group}/{detector}" for group, detector in missing[:20])
        ])

    return "\n".join(lines)


# ----------------------------- Bedienoberfläche -----------------------------

class Application(tk.Tk):
    def __init__(self) -> None:
        super().__init__()

        self.title("BMA Wartungsauswertung – Anmeldung")
        self.geometry("760x520")
        self.minsize(720, 500)
        self.configure(bg="#30363b")

        self.style = ttk.Style(self)
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass
        self.configure_styles()

        self.manufacturer_var = tk.StringVar(value="Hekatron")
        self.event_var = tk.StringVar()
        self.checklist_var = tk.StringVar()
        self.quarter_var = tk.StringVar(value="1")
        self.period_var = tk.StringVar(value="1 Tag")
        self.checklist_version_var = tk.StringVar(value="Version 1")
        self.esser_voralarm_var = tk.BooleanVar(value=False)
        self.create_txt_report_var = tk.BooleanVar(value=True)
        self.login_username_var = tk.StringVar()
        self.login_password_var = tk.StringVar()
        self.login_error_var = tk.StringVar()
        self.login_attempts = 0
        self.current_user = ""
        self.current_user_role = ""
        self.update_check_started = False

        self.status_var = tk.StringVar(value="Hersteller, Ereignisspeicher und Prüfliste auswählen.")
        self.stat_vars = {
            "q1": tk.StringVar(value="0"),
            "q2": tk.StringVar(value="0"),
            "q3": tk.StringVar(value="0"),
            "q4": tk.StringVar(value="0"),
            "total": tk.StringVar(value="0"),
            "open": tk.StringVar(value="0"),
            "checked": tk.StringVar(value="0"),
        }
        self.statistics_line_var = tk.StringVar(
            value="Gesamt: –  |  Offen: –  |  Q1: –  Q2: –  Q3: –  Q4: –"
        )
        self.traffic_light_var = tk.StringVar(value="● Keine Prüfliste ausgewählt")
        self.logo_images: Dict[str, tk.PhotoImage] = {}

        self.show_login()

    def configure_styles(self) -> None:
        """Modernes Grau-/Türkis-Design für Login und Hauptfenster."""
        turquoise = "#17b7b1"
        turquoise_dark = "#10948f"
        panel = "#3b4349"
        field = "#f3f5f6"

        self.style.configure("App.TFrame", background="#30363b")
        self.style.configure("Panel.TFrame", background=panel)
        self.style.configure("Card.TFrame", background="#414a50", relief="flat")
        self.style.configure("App.TLabel", background="#30363b", foreground="#f4f7f8")
        self.style.configure("Panel.TLabel", background=panel, foreground="#f4f7f8")
        self.style.configure("Title.TLabel", background="#30363b", foreground="#ffffff", font=("Segoe UI", 20, "bold"))
        self.style.configure("SubTitle.TLabel", background="#30363b", foreground="#aebbc1", font=("Segoe UI", 10))
        self.style.configure("Accent.TButton", background=turquoise, foreground="#ffffff", padding=(14, 9), font=("Segoe UI", 10, "bold"), borderwidth=0)
        self.style.map("Accent.TButton", background=[("active", turquoise_dark), ("pressed", turquoise_dark)])
        self.style.configure("Secondary.TButton", background="#59656c", foreground="#ffffff", padding=(10, 7), borderwidth=0)
        self.style.map("Secondary.TButton", background=[("active", "#68757c")])
        self.style.configure("Modern.TEntry", fieldbackground=field, foreground="#20262a", padding=7)
        self.style.configure("Modern.TCombobox", fieldbackground=field, foreground="#20262a", padding=5)
        self.style.configure("Modern.TCheckbutton", background=panel, foreground="#eef4f5", focuscolor=panel)
        self.style.map("Modern.TCheckbutton", background=[("active", panel)], foreground=[("disabled", "#889399")])

    def show_login(self) -> None:
        """Zeigt Benutzername und Passwort gemeinsam in einem modernen Loginfenster."""
        for child in self.winfo_children():
            child.destroy()

        self.title("BMA Wartungsauswertung – Anmeldung")
        self.geometry("760x520")
        self.configure(bg="#30363b")

        outer = ttk.Frame(self, style="App.TFrame", padding=34)
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        card = ttk.Frame(outer, style="Panel.TFrame", padding=36)
        card.grid(row=0, column=0, sticky="nsew", padx=70, pady=45)
        card.columnconfigure(0, weight=1)

        ttk.Label(card, text="BMA Wartungsauswertung", style="Panel.TLabel", font=("Segoe UI", 20, "bold")).grid(row=0, column=0, sticky="w")
        ttk.Label(card, text="Bitte mit einem berechtigten Benutzer anmelden", style="Panel.TLabel", foreground="#aebbc1").grid(row=1, column=0, sticky="w", pady=(5, 26))

        ttk.Label(card, text="Benutzername", style="Panel.TLabel").grid(row=2, column=0, sticky="w", pady=(0, 5))
        username_entry = ttk.Entry(card, textvariable=self.login_username_var, style="Modern.TEntry", font=("Segoe UI", 11))
        username_entry.grid(row=3, column=0, sticky="ew", ipady=4)

        ttk.Label(card, text="Passwort", style="Panel.TLabel").grid(row=4, column=0, sticky="w", pady=(16, 5))
        password_entry = ttk.Entry(card, textvariable=self.login_password_var, show="•", style="Modern.TEntry", font=("Segoe UI", 11))
        password_entry.grid(row=5, column=0, sticky="ew", ipady=4)

        ttk.Label(card, textvariable=self.login_error_var, style="Panel.TLabel", foreground="#ff8c8c").grid(row=6, column=0, sticky="w", pady=(9, 3))
        ttk.Button(card, text="Anmelden", command=self.attempt_login, style="Accent.TButton").grid(row=7, column=0, sticky="ew", pady=(12, 0), ipady=3)
        ttk.Label(card, text=f"Version: {VERSION}", style="Panel.TLabel", foreground="#87969d").grid(row=8, column=0, sticky="e", pady=(20, 0))

        username_entry.focus_set()
        username_entry.bind("<Return>", lambda _event: password_entry.focus_set())
        password_entry.bind("<Return>", lambda _event: self.attempt_login())

    def attempt_login(self) -> None:
        username = self.login_username_var.get().strip()
        password = self.login_password_var.get()

        try:
            role = verify_user(username, password)
        except Auswertungsfehler as exc:
            messagebox.showerror("Benutzerdatei", str(exc), parent=self)
            return

        if role is not None:
            self.current_user = username
            self.current_user_role = role
            self.login_error_var.set("")
            self.build_ui()
            return

        self.login_attempts += 1
        remaining = 3 - self.login_attempts
        self.login_password_var.set("")
        if remaining <= 0:
            messagebox.showerror("Zugriff verweigert", "Die Anmeldedaten wurden dreimal falsch eingegeben.", parent=self)
            self.destroy()
            return
        self.login_error_var.set(f"Benutzername oder Passwort falsch. Noch {remaining} Versuch(e).")

    def load_logo(self, filename: str) -> Optional[tk.PhotoImage]:
        """Lädt und verkleinert ein PNG/GIF neben dem Skript."""
        path = Path(__file__).resolve().parent / filename
        if not path.exists():
            return None

        try:
            image = tk.PhotoImage(file=str(path))
            width = max(1, image.width())
            height = max(1, image.height())

            # PhotoImage.subsample arbeitet mit ganzzahligen Faktoren.
            # Der größere notwendige Faktor begrenzt Breite und Höhe.
            factor_width = (width + MAX_LOGO_WIDTH - 1) // MAX_LOGO_WIDTH
            factor_height = (height + MAX_LOGO_HEIGHT - 1) // MAX_LOGO_HEIGHT
            factor = max(1, factor_width, factor_height)

            if factor > 1:
                image = image.subsample(factor, factor)
            return image
        except tk.TclError:
            return None

    def select_manufacturer(self, manufacturer: str) -> None:
        self.manufacturer_var.set(manufacturer)
        if manufacturer == "Esser":
            self.esser_voralarm_checkbox.configure(state="normal")
            self.status_var.set(
                "Esser ausgewählt. Optional können Voralarm-Meldungen mit ausgewertet werden."
            )
        else:
            self.esser_voralarm_var.set(False)
            self.esser_voralarm_checkbox.configure(state="disabled")
            self.status_var.set(
                "Hekatron ausgewählt. Bitte Ereignisspeicher und Prüfliste auswählen."
            )

    def build_manufacturer_card(
        self,
        parent: ttk.Frame,
        manufacturer: str,
        logo_filename: str,
        column: int,
    ) -> None:
        card = ttk.Frame(parent, padding=12, style="Card.TFrame")
        card.grid(row=0, column=column, sticky="nsew", padx=8)
        card.columnconfigure(0, weight=1)

        image = self.load_logo(logo_filename)
        if image is not None:
            self.logo_images[manufacturer] = image
            logo = ttk.Label(card, image=image, anchor="center")
            logo.grid(row=0, column=0, pady=(2, 8))

        ttk.Radiobutton(
            card,
            text=manufacturer,
            variable=self.manufacturer_var,
            value=manufacturer,
            command=lambda: self.select_manufacturer(manufacturer),
        ).grid(row=1, column=0, pady=(4, 2))

        ttk.Button(
            card,
            text=f"{manufacturer} auswählen",
            command=lambda: self.select_manufacturer(manufacturer),
        ).grid(row=2, column=0, sticky="ew", pady=(8, 0))

    def build_ui(self) -> None:
        for child in self.winfo_children():
            child.destroy()

        self.title(f"BMA Wartungsauswertung – {self.current_user}")
        self.geometry("980x760")
        self.minsize(900, 700)
        self.configure(bg="#30363b")

        frame = ttk.Frame(self, style="App.TFrame", padding=22)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="BMA Wartungsauswertung", style="Title.TLabel").grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(frame, text=f"Angemeldet als {self.current_user} ({self.current_user_role})", style="SubTitle.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", pady=(2, 16))
        ttk.Label(frame, text=f"Version: {VERSION}", style="SubTitle.TLabel").grid(row=0, column=2, rowspan=2, sticky="ne")

        manufacturer_frame = ttk.Frame(frame, style="Panel.TFrame", padding=14)
        manufacturer_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 14))
        manufacturer_frame.columnconfigure(0, weight=1)
        manufacturer_frame.columnconfigure(1, weight=1)
        self.build_manufacturer_card(manufacturer_frame, "Hekatron", "hekatron_logo.png", 0)
        self.build_manufacturer_card(manufacturer_frame, "Esser", "esser_logo.png", 1)

        form = ttk.Frame(frame, style="Panel.TFrame", padding=18)
        form.grid(row=3, column=0, columnspan=3, sticky="nsew")
        form.columnconfigure(1, weight=1)

        ttk.Label(form, text="Ereignisspeicher", style="Panel.TLabel").grid(row=0, column=0, sticky="w", pady=7)
        ttk.Entry(form, textvariable=self.event_var, style="Modern.TEntry").grid(row=0, column=1, sticky="ew", padx=10, pady=7)
        ttk.Button(form, text="Auswählen …", command=self.choose_event, style="Secondary.TButton").grid(row=0, column=2, pady=7)

        ttk.Label(form, text="Prüfliste (ODS/Excel)", style="Panel.TLabel").grid(row=1, column=0, sticky="w", pady=7)
        ttk.Entry(form, textvariable=self.checklist_var, style="Modern.TEntry").grid(row=1, column=1, sticky="ew", padx=10, pady=7)
        ttk.Button(form, text="Auswählen …", command=self.choose_checklist, style="Secondary.TButton").grid(row=1, column=2, pady=7)

        ttk.Label(form, text="Prüflisten-Version", style="Panel.TLabel").grid(row=2, column=0, sticky="w", pady=7)
        version_box = ttk.Combobox(form, textvariable=self.checklist_version_var, values=("Version 1", "Version 2"), state="readonly", width=16, style="Modern.TCombobox")
        version_box.grid(row=2, column=1, sticky="w", padx=10, pady=7)
        version_box.bind("<<ComboboxSelected>>", lambda _event: self.update_statistics())

        ttk.Label(form, text="Quartal", style="Panel.TLabel").grid(row=3, column=0, sticky="w", pady=7)
        ttk.Combobox(form, textvariable=self.quarter_var, values=("1", "2", "3", "4"), state="readonly", width=10, style="Modern.TCombobox").grid(row=3, column=1, sticky="w", padx=10, pady=7)

        ttk.Label(form, text="Auswertungszeitraum", style="Panel.TLabel").grid(row=4, column=0, sticky="w", pady=7)
        ttk.Combobox(form, textvariable=self.period_var, values=("1 Tag", "3 Tage", "7 Tage", "30 Tage"), state="readonly", width=14, style="Modern.TCombobox").grid(row=4, column=1, sticky="w", padx=10, pady=7)

        self.esser_voralarm_checkbox = ttk.Checkbutton(form, text="Bei Esser auch VORALARM und VORALARM-ENDE berücksichtigen", variable=self.esser_voralarm_var, state="disabled", style="Modern.TCheckbutton")
        self.esser_voralarm_checkbox.grid(row=5, column=0, columnspan=3, sticky="w", pady=(9, 4))

        ttk.Checkbutton(form, text="TXT-Prüfbericht mit offenen Meldern erstellen", variable=self.create_txt_report_var, style="Modern.TCheckbutton").grid(row=6, column=0, columnspan=3, sticky="w", pady=(4, 10))

        stats_frame = ttk.Frame(frame, style="Panel.TFrame", padding=(12, 9))
        stats_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(14, 10))
        stats_frame.columnconfigure(0, weight=1)
        ttk.Label(stats_frame, textvariable=self.statistics_line_var, style="Panel.TLabel", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.traffic_light_label = tk.Label(stats_frame, textvariable=self.traffic_light_var, fg="#f4c95d", bg="#3b4349", font=("Segoe UI", 9), anchor="e")
        self.traffic_light_label.grid(row=0, column=1, sticky="e", padx=(12, 0))

        ttk.Button(frame, text="Prüfliste ausfüllen und exportieren", command=self.export, style="Accent.TButton").grid(row=5, column=0, columnspan=3, sticky="ew", pady=(4, 12), ipady=5)

        ttk.Label(frame, textvariable=self.status_var, style="App.TLabel", wraplength=790, justify="left").grid(row=6, column=0, columnspan=2, sticky="w", pady=7)
        action_frame = ttk.Frame(frame, style="App.TFrame")
        action_frame.grid(row=6, column=2, sticky="se", padx=(10, 0), pady=7)
        ttk.Button(action_frame, text="Update prüfen", command=lambda: self.start_update_check(True), style="Secondary.TButton").pack(side="left", padx=(0, 6))
        ttk.Button(action_frame, text="Hilfe", command=self.show_help, style="Secondary.TButton", width=10).pack(side="left")

        self.status_var.set(f"Angemeldet als: {self.current_user} ({self.current_user_role}). Bitte Dateien auswählen.")

        # Updateprüfung läuft im Hintergrund und blockiert den Programmstart nicht.
        if not self.update_check_started:
            self.update_check_started = True
            self.after(1200, self.start_update_check)

    def start_update_check(self, manual: bool = False) -> None:
        """Startet die Versionsprüfung ohne das Fenster einzufrieren."""
        def worker() -> None:
            try:
                data = load_online_version()
                self.after(0, lambda: self.handle_update_result(data, manual))
            except Exception as exc:
                if manual:
                    self.after(0, lambda: messagebox.showwarning(
                        "Versionsprüfung",
                        "Die Online-Version konnte nicht geprüft werden.\n\n"
                        f"Fehler: {exc}",
                        parent=self,
                    ))
                else:
                    print(f"Versionsprüfung nicht möglich: {exc}")

        threading.Thread(target=worker, daemon=True).start()

    def handle_update_result(self, data: Dict[str, Any], manual: bool = False) -> None:
        """Vergleicht die Versionen und bietet ein echtes Ein-Klick-Update an."""
        try:
            online_version = str(data.get("version", "")).strip()
            if parse_version(online_version) <= parse_version(VERSION):
                if manual:
                    messagebox.showinfo(
                        "Versionsprüfung",
                        f"Du verwendest bereits die aktuelle Version {VERSION}.",
                        parent=self,
                    )
                return

            titel = str(data.get("titel", "Update verfügbar")).strip() or "Update verfügbar"
            hinweis = str(data.get("hinweis", "")).strip()
            download_url = str(data.get("download_url", "")).strip()
            expected_sha256 = str(data.get("sha256", "")).strip().lower()
            pflichtupdate = bool(data.get("pflichtupdate", False))

            text = (
                f"Eine neue Version ist verfügbar.\n\n"
                f"Installiert: {VERSION}\n"
                f"Verfügbar: {online_version}"
            )
            if hinweis:
                text += f"\n\nÄnderungen:\n{hinweis}"

            if not download_url:
                messagebox.showinfo(
                    titel,
                    text + "\n\nFür dieses Update ist noch kein Download hinterlegt.",
                    parent=self,
                )
                return

            text += "\n\nJetzt automatisch aktualisieren?"
            if messagebox.askyesno(titel, text, parent=self):
                self.download_and_install_update(
                    download_url=download_url,
                    online_version=online_version,
                    expected_sha256=expected_sha256,
                )
            elif pflichtupdate:
                self.status_var.set(f"Pflichtupdate verfügbar: Version {online_version}")

        except Exception as exc:
            if manual:
                messagebox.showerror(
                    "Versionsprüfung",
                    f"Die Versionsangaben konnten nicht verarbeitet werden.\n\n{exc}",
                    parent=self,
                )

    def download_and_install_update(
        self,
        download_url: str,
        online_version: str,
        expected_sha256: str = "",
    ) -> None:
        """Lädt die neue PY-Datei, ersetzt die laufende Datei und startet neu."""
        self.status_var.set(f"Update {online_version} wird heruntergeladen …")
        self.title("BMA Wartungsauswertung – Update wird installiert …")
        self.update_idletasks()

        def worker() -> None:
            temp_path: Optional[Path] = None
            try:
                request = urllib.request.Request(
                    download_url,
                    headers={
                        "User-Agent": "BMA-Wartungsauswertung",
                        "Cache-Control": "no-cache",
                    },
                )
                with urllib.request.urlopen(
                    request,
                    timeout=30,
                    context=create_ssl_context(),
                ) as response:
                    content = response.read()

                if len(content) < 1000:
                    raise Auswertungsfehler(
                        "Die heruntergeladene Programmdatei ist ungewöhnlich klein."
                    )

                try:
                    source = content.decode("utf-8-sig")
                except UnicodeDecodeError as exc:
                    raise Auswertungsfehler(
                        "Die heruntergeladene Datei ist keine gültige Python-Datei."
                    ) from exc

                version_match = re.search(
                    r'^VERSION\s*=\s*["\']([^"\']+)["\']',
                    source,
                    flags=re.MULTILINE,
                )
                if not version_match:
                    raise Auswertungsfehler(
                        "In der heruntergeladenen Datei wurde keine Versionsnummer gefunden."
                    )
                downloaded_version = version_match.group(1).strip()
                if parse_version(downloaded_version) != parse_version(online_version):
                    raise Auswertungsfehler(
                        "Die heruntergeladene Programmversion stimmt nicht mit "
                        "der angekündigten Version überein."
                    )

                if expected_sha256:
                    actual_sha256 = hashlib.sha256(content).hexdigest().lower()
                    if actual_sha256 != expected_sha256:
                        raise Auswertungsfehler(
                            "Die Prüfsumme des Updates stimmt nicht. Das Update wurde abgebrochen."
                        )

                compile(source, "BMA_Wartungsauswertung.py", "exec")

                current_file = Path(__file__).resolve()
                backup_file = current_file.with_suffix(".backup.py")
                fd, temp_name = tempfile.mkstemp(
                    prefix="BMA_Update_",
                    suffix=".py",
                    dir=str(current_file.parent),
                )
                os.close(fd)
                temp_path = Path(temp_name)
                temp_path.write_bytes(content)

                if backup_file.exists():
                    backup_file.unlink()
                current_file.replace(backup_file)
                try:
                    temp_path.replace(current_file)
                except Exception:
                    if not current_file.exists() and backup_file.exists():
                        backup_file.replace(current_file)
                    raise

                self.after(0, lambda: self.finish_update_and_restart(online_version, current_file))

            except Exception as exc:
                if temp_path is not None and temp_path.exists():
                    try:
                        temp_path.unlink()
                    except OSError:
                        pass
                self.after(0, lambda error=exc: self.update_failed(error))

        threading.Thread(target=worker, daemon=True).start()

    def finish_update_and_restart(self, online_version: str, current_file: Path) -> None:
        messagebox.showinfo(
            "Update abgeschlossen",
            f"Version {online_version} wurde erfolgreich installiert.\n\n"
            "Das Programm wird jetzt neu gestartet.",
            parent=self,
        )
        os.execv(sys.executable, [sys.executable, str(current_file)])

    def update_failed(self, error: Exception) -> None:
        self.title("BMA Wartungsauswertung")
        self.status_var.set("Update fehlgeschlagen. Das Programm kann weiter benutzt werden.")
        messagebox.showerror(
            "Update fehlgeschlagen",
            "Das Update konnte nicht installiert werden.\n\n"
            f"Fehler: {error}",
            parent=self,
        )

    def show_help(self) -> None:
        """Zeigt vorläufig einen Platzhalter für die späteren Anleitungen."""
        messagebox.showinfo(
            "Hilfe",
            "Ich arbeite dran",
            parent=self,
        )

    def choose_event(self) -> None:
        manufacturer = self.manufacturer_var.get()
        if manufacturer == "Hekatron":
            filetypes = [
                ("Hekatron-Ereignisspeicher (Excel)", "*.xlsx"),
                ("Excel-Dateien", "*.xlsx"),
                ("Alle Dateien", "*.*"),
            ]
        else:
            filetypes = [
                ("Esser-Ereignisspeicher (PDF)", "*.pdf"),
                ("PDF-Dateien", "*.pdf"),
                ("Alle Dateien", "*.*"),
            ]

        path = filedialog.askopenfilename(
            title=f"{manufacturer}-Ereignisspeicher auswählen",
            filetypes=filetypes,
        )
        if path:
            self.event_var.set(path)

    def choose_checklist(self) -> None:
        path = filedialog.askopenfilename(
            title="Prüfliste auswählen",
            filetypes=[
                ("Prüflisten (ODS oder Excel)", "*.ods *.xlsx *.xls"),
                ("OpenDocument-Tabelle", "*.ods"),
                ("Excel-Arbeitsmappe", "*.xlsx"),
                ("Excel 97–2003-Arbeitsmappe", "*.xls"),
                ("Alle Dateien", "*.*"),
            ],
        )
        if path:
            self.checklist_var.set(path)
            self.update_statistics()

    def update_statistics(self, path: Optional[str] = None) -> None:
        checklist_path = (path or self.checklist_var.get()).strip()
        if not checklist_path:
            return
        try:
            stats = prueflisten_statistik(checklist_path, self.checklist_version_var.get())
        except Auswertungsfehler as exc:
            self.status_var.set(f"Statistik konnte nicht geladen werden: {exc}")
            return

        for key, value in stats.items():
            self.stat_vars[key].set(str(value))
        checked = max(0, stats["total"] - stats["open"])
        self.stat_vars["checked"].set(str(checked))
        self.statistics_line_var.set(
            f"Gesamt: {stats['total']}  |  Offen: {stats['open']}  |  "
            f"Q1: {stats['q1']}  Q2: {stats['q2']}  Q3: {stats['q3']}  Q4: {stats['q4']}"
        )
        if stats["total"] <= 0:
            self.traffic_light_var.set("● Keine Melder gefunden")
            self.traffic_light_label.configure(fg="#b91c1c")
        elif stats["open"] == 0:
            self.traffic_light_var.set("● Alle Melder geprüft")
            self.traffic_light_label.configure(fg="#15803d")
        else:
            self.traffic_light_var.set(f"● Noch {stats['open']} offen")
            self.traffic_light_label.configure(fg="#ca8a04")
        self.status_var.set(
            f"Statistik geladen: {stats['total']} Melder insgesamt, "
            f"{checked} geprüft und {stats['open']} noch offen."
        )

    def export(self) -> None:
        manufacturer = self.manufacturer_var.get()
        event_path = self.event_var.get().strip()
        checklist_path = self.checklist_var.get().strip()
        if not event_path or not checklist_path:
            messagebox.showwarning(
                "Dateien fehlen",
                "Bitte Ereignisspeicher und Prüfliste auswählen.",
            )
            return

        original = Path(checklist_path)
        extension = original.suffix.lower()
        if extension == ".xlsx":
            save_types = [("Excel-Arbeitsmappe", "*.xlsx")]
        elif extension == ".xls":
            save_types = [("Excel 97–2003-Arbeitsmappe", "*.xls")]
        else:
            extension = ".ods"
            save_types = [("OpenDocument-Tabelle", "*.ods")]

        output = filedialog.asksaveasfilename(
            title="Ausgefüllte Prüfliste speichern",
            defaultextension=extension,
            initialfile=f"{original.stem}_Q{self.quarter_var.get()}_ausgefüllt{extension}",
            filetypes=save_types,
        )
        if not output:
            return

        self.status_var.set("Auswertung läuft …")
        self.title("BMA Wartungsauswertung – Verarbeite Ereignisse …")
        self.update_idletasks()

        try:
            quarter = int(self.quarter_var.get())
            zeitraum_tage = int(self.period_var.get().split()[0])
            result = pruefliste_ausfuellen(
                event_path=event_path,
                checklist_path=checklist_path,
                output_path=output,
                quarter=quarter,
                zeitraum_tage=zeitraum_tage,
                manufacturer=manufacturer,
                esser_voralarm=(manufacturer == "Esser" and self.esser_voralarm_var.get()),
                checklist_version=self.checklist_version_var.get(),
            )
            if self.create_txt_report_var.get():
                result["report"] = txt_pruefbericht_schreiben(result, quarter)
            else:
                result["report"] = None
            self.update_statistics(str(result["output"]))
        except Auswertungsfehler as exc:
            self.status_var.set(f"Fehler: {exc}")
            self.title("BMA Wartungsauswertung – Fehler")
            messagebox.showerror("Auswertung fehlgeschlagen", str(exc))
            return
        except Exception as exc:
            self.status_var.set(f"Unerwarteter Fehler: {exc}")
            self.title("BMA Wartungsauswertung – Fehler")
            messagebox.showerror(
                "Unerwarteter Fehler", f"{type(exc).__name__}: {exc}"
            )
            return

        text = result_text(result)
        self.status_var.set(text)
        self.title("BMA Wartungsauswertung – Fertig ✓")
        messagebox.showinfo("Auswertung abgeschlossen", text)

def main() -> None:
    app = Application()
    app.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        error_path = Path(__file__).resolve().parent / "BMA_Fehlerprotokoll.txt"
        error_path.write_text(traceback.format_exc(), encoding="utf-8")
        raise
